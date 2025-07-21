
from AgenciamientoAduanal.models import AgenteAduanal
from Clientes.mixins import ClienteRolMixin
from Core.mixins import N4GroovyMixin
from Solicitudes.models import SolicitudIndividual
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import os
from django.http import HttpResponse
from django.conf import settings
from Solicitudes.models import SolicitudIndividual
from django.contrib.auth.models import User


class SegregarContenedores(N4GroovyMixin, ClienteRolMixin):
    
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)

        def Segregar(self,solicitud: SolicitudIndividual):

            if self.es_cliente_padre(self.request.user):
                RazonSocial = solicitud.solicitud_masiva.cliente_padre.razon_social
            else:
                RazonSocial = solicitud.solicitud_masiva.cliente_hijo.cliente_padre.razon_social

            IdParametros = ["Contenedor", "Grupo", "TipoOperacion", "NombreCliente"]
            Parametros = [ solicitud.contenedor
                          ,solicitud.solicitud_masiva.tipo_segregacion.grupo
                          ,solicitud.solicitud_masiva.tipo_segregacion.tipo_operacion
                          ,RazonSocial]

            self.Cliente = self.Conexion()
            Error = self.EjecutarGroovy(IdParametros,Parametros, "GvyLogiPuertoSegregacion")

            #EJEMPLO DE EJECUCION DEL GROOVY
            # <groovy class-name="GvyLogiPuertoSegregacion" class-location="database">
            # <parameters>
            # <parameter id="Contenedor" value="FCIU9159462"/>
            # <parameter id="Grupo" value="LOG"/>
            # <parameter id="TipoOperacion" value="LOGIPUERTO"/>
            # <parameter id="NombreCliente" value="PRUEBA"/>
            # </parameters>
            # </groovy>

            # Si no hubo un error al ejecutar el groovy se realiza el cambio en bd

            if not Error.get("Error"):
                solicitud.estatus_n4 = "Autorizado"
                solicitud.save()

        def ActualizarContenedor(self, solicitud: SolicitudIndividual, xml_data: dict):
            # <column>Gkey</column>
            # <column>BL</column>
            # <column>Type Arch ISO</column>
            # <column>Buque</column>
            # <column>Contenedor</column>
            # <column>Reefer</column>
            # <column>ETA</column>
            # <column>IMO</column>
            # <column>OOG</column>
            # <column>Viaje</column>
            solicitud.id_contenedor = xml_data.get('Gkey', '')
            solicitud.bl = xml_data.get('BL', '')           
            solicitud.tipo_tamaño = xml_data.get('Type Arch ISO', '')
            solicitud.buque = xml_data.get('Buque', '')
            solicitud.esReefer = False if xml_data.get('Reefer') in [None, "NON_RFR"] else True
 
            if  xml_data.get('ETA', '') == "None":
                solicitud.eta = None
            else:
                solicitud.eta = xml_data.get('ETA', '')            
            
            if xml_data.get('IMO', '') == "None":
                solicitud.imo = None
            else:
                solicitud.imo = bool(xml_data.get('IMO'))
            
            solicitud.isOOG = xml_data.get('OOG', '')           
            solicitud.viaje = xml_data.get('Viaje', '')           
            solicitud.estado = "Actualizado"
            solicitud.save()

        
        def eliminar_grupo_tipo_N4(self,contenedor):
            self.Cliente = self.Conexion()
            respuesta = self.EjecutarGroovy(["Contenedor"], [contenedor], "GvyLogCancelarSol")            
            return respuesta
            #Ejemplo de groovy a ejecutar
            #<groovy class-name="GvyLogCancelarSoli" class-location="database">
            #   <parameters>
            #       <parameter id="Contenedor" value="CSNU2027309"/>
            #   </parameters>
            #</groovy>


        def eliminar_solicitud(self, pk):
            solicitud = SolicitudIndividual.objects.get(pk = pk)
            solicitud.eliminada = True
            solicitud.save()

class ControladorExcel():       
    def editar_excel(self, Archivo, pk):
        # Cargar archivo existente
        ruta_archivo = os.path.join(settings.MEDIA_ROOT, Archivo)
        workbook = load_workbook(ruta_archivo)
        worksheet = workbook.active
        
        if not os.path.exists(ruta_archivo):
            raise FileNotFoundError(f"No se encontró el archivo en: {ruta_archivo}")
        
        # Headers
        worksheet['B1'] = 'Tipo'
        worksheet['C1'] = 'BL'
        worksheet['D1'] = 'Destino'
        worksheet['E1'] = 'Viaje'
        worksheet['F1'] = 'Buque'
        worksheet['G1'] = 'ETA'
        worksheet['H1'] = 'IMO'
        worksheet['I1'] = 'EsReeferr'
        worksheet['J1'] = 'IsOOG'        
        
        # Obtener todas las solicitudes 
        solicitudes = SolicitudIndividual.objects.filter(solicitud_masiva=pk)
        fila = 2  # Empezar en fila 2 porque la 1 tiene headers
        
        for solicitud in solicitudes:
            worksheet.cell(row=fila, column=2, value=solicitud.tipo_tamaño)  # B
            worksheet.cell(row=fila, column=3, value=solicitud.bl)           # C
            worksheet.cell(row=fila, column=4, value=solicitud.destino)     # D
            worksheet.cell(row=fila, column=5, value=solicitud.viaje)        # E
            worksheet.cell(row=fila, column=6, value=solicitud.buque)        # F
            worksheet.cell(row=fila, column=7, value=solicitud.eta)          # G
            worksheet.cell(row=fila, column=8, value=solicitud.imo)          # H
            worksheet.cell(row=fila, column=9, value=solicitud.esReefer)     # I
            worksheet.cell(row=fila, column=10, value=solicitud.isOOG)        # J
            fila += 1  
        
        workbook.save(ruta_archivo)
        workbook.close()

class ValidacionesTransporte():

    def comparar_patente_n4(self, usuario: User, xml_data: dict):
        # <column>Patente</column>
        patente = xml_data.get('Patente', '').strip()
        
        agencia_aduanal = usuario.usuario_agente_aduanal.agencia_aduanal

        # Obtener lista de patentes de la agencia
        patentes = AgenteAduanal.objects.filter(
            agencia_aduanal=agencia_aduanal,
            patente__isnull=False
        ).values_list('patente', flat=True)
        print("-----------------------------------------")
        print(patentes)

        return patente in patentes



        
        
