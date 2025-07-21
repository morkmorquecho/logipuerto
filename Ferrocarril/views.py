from datetime import date, datetime, timedelta, timezone
from urllib.parse import urlencode
from django.utils import timezone
from django.core.files.base import File
import os
import tempfile
import time
from django.forms import modelformset_factory
from typing import Any
from django.db import DatabaseError
from django.forms import ValidationError
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.views.generic import CreateView, TemplateView, ListView, FormView
from django.db.models import Q
from django.core.files.storage import default_storage
import openpyxl
import requests
import sentry_sdk
from Core.mixins import ErrorHandlingMixin, GroupRequiredMixin, MessageMixin
from Ferrocarril.models import SoliciudIndividualTransporte
from Solicitudes.service import verificar_contenedor
from Ferrocarril.form import BuscarContenedoresPorBLForm, CrearSolicitudesTransporteForm, SolicitudesMasivasFerroForm
from Clientes.models import ClienteHijo
from Solicitudes.models import SolicitudMasiva, SolicitudIndividual, ClientePadre
from Clientes.models import TipoDeServicio
from Solicitudes.models import TipoSegregacion
from Solicitudes.mixins import ConsultarPreAvisoMixin 
from Ferrocarril.service import SegregarContenedores, ControladorExcel, ValidacionesTransporte
from django.http import JsonResponse
from django.views import View

class FerrocarrilView(GroupRequiredMixin, TemplateView):
    template_name = "Ferrocarril/inicio.html"
    group_required = 'ClienteHijo', 'ClientePadre'


class SegregacionView(GroupRequiredMixin, MessageMixin, CreateView, ConsultarPreAvisoMixin, SegregarContenedores, ControladorExcel):
    model = SolicitudMasiva
    form_class = SolicitudesMasivasFerroForm
    template_name = "Ferrocarril/segregacion_form.html"
    group_required = 'ClienteHijo', 'ClientePadre'

    def get_success_url(self):
        return reverse_lazy('solicitudes:solicitudes_masivas__cliente_padre', kwargs={'pk': self.request.user.id})

    def form_valid(self, form):
        user = self.request.user
        #Ingrar datos para la solicitud Masiva:
        solicitudMasiva = form.save(commit=False)
        #Consultar si estamos ante un cliente hijo
        cliente_hijo = ClienteHijo.objects.filter(user=user).first()
        digito_verificador = self.registrar_cliente(solicitudMasiva, cliente_hijo, user)
        

        solicitudMasiva.user = user
        solicitudMasiva.servicio = form.cleaned_data['tipo_segregacion'].tipo_servicio

        if 'excel' not in self.request.FILES:
            form.add_error('excel', 'Debe subir un archivo Excel')
            return self.form_invalid(form)

        #Empezamos a procesar el excel:
        #Creamos el temporal para manejar el excel
        tmp_path = self.crear_archivo_tmp(self.request.FILES['excel'])
        solicitudes = None
        errores = None
        try:
            #obtenemos las solicitudes del excel dado
            solicitudes = self.procesar_excel(tmp_path)
            #Si el digito verificador esta activo validamos cada contenedor y obtenemos errores si los hay
            if digito_verificador:
                errores = self.validar_contenedores(solicitudes)
                if errores:
                    self.error(f"Se encontraron los siguientes errores: \n{errores}")
                    return self.form_invalid(form)
        except Exception as e:
            self.error("Error al procesar el archivo Excel")
            print("Error al procesar el archivo Excel")
            print(e)
            # Enviamos el error a Sentry
            with sentry_sdk.push_scope() as scope:
                scope.set_tag("Aplicacion", "Logipuerto")
                scope.set_tag("Modulo", "SegregacionView")
                scope.set_tag("Proceso", "Validacion de digito verificado")
                scope.set_tag("digito_verificador", digito_verificador)
                scope.set_context("solicitudes", solicitudes)
                scope.set_context("errores detectados en el excel", errores)
                sentry_sdk.capture_exception(e)                    

            return self.form_invalid(form)
        
        try:
            self.guardar_solicitud_masiva(solicitudMasiva, tmp_path)
            self.crear_solicitudes_individuales(solicitudMasiva, solicitudes)
            self.actualizar_excel(solicitudMasiva)
        except Exception as e:
            self.error("Ocurrió un error procesando la solicitud, de continua con este problema contacte soporte")
            print("Error al guardar la solicitud masiva o las solicitudes individuales")
            print(e)
            with sentry_sdk.push_scope() as scope:
                scope.set_tag("Aplicacion", "Logipuerto")
                scope.set_tag("Modulo", "SegregacionView")
                scope.set_tag("Proceso", "Actualizar excel, segregacion en N4, Actualizacion de solicitudes individuales")
                scope.set_context("solicitudes", solicitudes)
                scope.set_context("solicitudMasiva", solicitudMasiva)
                sentry_sdk.capture_exception(e)   
            return self.form_invalid(form)

        return super().form_valid(form)

    def registrar_cliente(self, solicitudMasiva, cliente_hijo, user):
        if cliente_hijo:
            solicitudMasiva.cliente_hijo = cliente_hijo
            solicitudMasiva.cliente_padre = cliente_hijo.cliente_padre
            return cliente_hijo.cliente_padre.digito_verificador
        else:
            cliente_padre = ClientePadre.objects.filter(user=user).first()
            print(f"Cliente padre-------------------: {cliente_padre}")
            if cliente_padre:
                solicitudMasiva.cliente_padre = cliente_padre
                return cliente_padre.digito_verificador
            else:
                raise ValueError("El usuario no tiene asignado ni ClienteHijo ni ClientePadre")
            
    def crear_archivo_tmp(self, excel):
            # Guardamos el archivo en un temporal
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                for chunk in excel.chunks():
                    tmp.write(chunk)
                    tmp_path = tmp.name
            return tmp_path

    def procesar_excel(self, tmp_path):
        workbook = openpyxl.load_workbook(tmp_path, read_only=True)
        sheet = workbook.active
        solicitudes = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4): 
            solicitudes.append({
                "contenedor": row[0].value,
                "tipo tamaño": row[1].value,
                "BL": row[2].value,
                "destino": row[3].value,
            })
        workbook.close()
        return solicitudes

    def validar_contenedores(self, solicitudes):
        errores = ""
        fila = 1
        for solicitud in solicitudes:
            fila += 1
            contenedor = solicitud.get("contenedor")
            numero, valido = verificar_contenedor(self.request, contenedor)
            if not valido:
                errores += f"Contenedor inválido: ({numero}) fila {fila} - \n"
        return errores

    def guardar_solicitud_masiva(self, solicitudMasiva, tmp_path):
        solicitudMasiva.save()
        NombreArchivo = f"SolicitudMasiva{solicitudMasiva.id}"
        with open(tmp_path, 'rb') as f:
            solicitudMasiva.excel.save(f'{NombreArchivo}.xlsx', File(f), save=False)
        os.unlink(tmp_path)
        solicitudMasiva.save()


    def crear_solicitudes_individuales(self, solicitudMasiva, solicitudes):
        for solicitud in solicitudes:
            SolicitudIndividual.objects.create(
                solicitud_masiva=solicitudMasiva,
                bl=solicitud.get('BL'),
                contenedor=solicitud.get('contenedor'),
                tipo_tamaño=solicitud.get('tipo tamaño'),
                destino=solicitud.get('destino')
            )

    def actualizar_excel(self, solicitudMasiva):
        solicitudes = SolicitudIndividual.objects.filter(solicitud_masiva=solicitudMasiva.pk)
        
        for solicitud in solicitudes:
            succes, datos = self.Consultar(solicitud)
            if succes:
                self.ActualizarContenedor(solicitud, datos)
                self.Segregar(solicitud)
                self.editar_excel(solicitud.solicitud_masiva.excel.name, solicitud.solicitud_masiva)

            
    def form_invalid(self, form):
        print(f"error del formulario", form.errors)
        print(form.errors.as_json())
        return super().form_invalid(form)
      
class SolicitudesIndividualesFerroView(ErrorHandlingMixin,GroupRequiredMixin, ListView, ConsultarPreAvisoMixin, SegregarContenedores, ControladorExcel):
    def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)

    model = SolicitudIndividual
    template_name = "Ferrocarril/SolicitudIndividuales_list.html"
    context_object_name = 'solicitudes_individuales'
    paginate_by = 20
    ordering = ['-fecha_creacion']
    group_required = 'ClienteHijo', 'Logipuerto', 'ClientePadre' 

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        solicitudes = self.get_queryset()
        hoy = timezone.now()
        fecha_limite = hoy - timedelta(days=30)

        for solicitud in solicitudes:
            if solicitud.estado == "Nueva" or solicitud.estatus_n4 == "Nueva":
                if solicitud.fecha_creacion < fecha_limite:
                    Error = self.eliminar_grupo_tipo_N4(solicitud.contenedor)
                    if not Error:
                        self.eliminar_solicitud(solicitud.pk)

                succes, datos = self.Consultar(solicitud)
                if succes:
                    if solicitud.estatus_n4 == "Nueva":
                        self.Segregar(solicitud)

                    if solicitud.estado == "Nueva":
                        self.ActualizarContenedor(solicitud, datos)
                        self.editar_excel(solicitud.solicitud_masiva.excel.name ,solicitud.solicitud_masiva)

        return context
    

    def get_queryset(self):
        queryset = super().get_queryset().order_by('-fecha_creacion')
        search_query = self.request.GET.get('q')

        Estado = self.request.GET.get('Estado')
        if Estado:
            queryset = queryset.filter(estado=Estado)

        # Barra de búsqueda
        if search_query:
                try:
                    queryset = queryset.filter(
                        Q(bl=search_query) |  
                        Q(contenedor=search_query) |
                        Q(id=search_query))
                except ValueError:
                    # Si falla por el id no numérico, filtramos solo por bl y contenedor
                    queryset = queryset.filter(
                        Q(bl=search_query) |  
                        Q(contenedor=search_query))
    
        return queryset.filter(solicitud_masiva = self.kwargs['pk'], eliminada = False)

class FerrocarrilTransporteView(GroupRequiredMixin, FormView, ConsultarPreAvisoMixin, MessageMixin):
    template_name = "Ferrocarril/form_get.html"
    group_required = 'AgenteAduanal'
    form_class = BuscarContenedoresPorBLForm

    def get_context_data(self, **kwargs):
        return super().get_context_data(**kwargs)

        
    def form_valid(self, form):
        bl = form.cleaned_data['bl']
        user_id = self.request.user
        solicitudes = SolicitudIndividual.objects.filter(bl = bl, estado = "Actualizado", estatus_n4 = "Autorizado")
        print(solicitudes)
        if solicitudes == None:
            self.error("No existen contenedores con ese bl en el sistema")

        if not self.validar_patente(user_id, solicitudes):
            self.error("No cuentas con la patente necesaria")
            return self.form_invalid(form)

        url = reverse(
            'ferrocarril:ferrocarril_transporte_escoger_contenedor'
        )

        query_string = urlencode({'bl': bl})
        return redirect(f"{url}?{query_string}")
    
    def validar_patente(self, user_id, solicitudes):
        validacionesTransporte = ValidacionesTransporte
        for solicitud in solicitudes:
            succes, datos = self.Consultar(solicitud)
            if succes:
                if not validacionesTransporte.comparar_patente_n4(self, user_id, datos):
                    # self.error("No cuentas con la patente necesaria")
                    return False #No existe patente
        return True #Existe Patente

                    
class EscogerContenedoresView(GroupRequiredMixin, TemplateView):
    template_name = "Ferrocarril/escoger_contenedor.html"
    group_required = 'AgenteAduanal'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        bl = self.request.GET.get('bl')
        
        solicitudes = SolicitudIndividual.objects.filter(bl = bl, estado = "Actualizado", estatus_n4 = "Autorizado")

        context.update({
            'bl': bl,
            'contenedores': solicitudes.values_list('contenedor', flat=True)
        })
        return context
    
    def post(self, request, *args, **kwargs):
        contenedores_seleccionados = request.POST.getlist('contenedores_seleccionados')
        bl = request.POST.get('bl') or request.GET.get('bl')
        
        # Guardar en sesión
        request.session['contenedores_seleccionados'] = contenedores_seleccionados
        request.session['bl'] = bl
        
        return redirect('ferrocarril:ferrocarril_transporte_crear_solicitud_masiva')


        # request.session.pop('contenedores_seleccionados', None)
        # request.session.pop('bl', None)
        # request.session.pop('timestamp_seleccion', None)


class CrearSolicitudesIndividualesTransporteView(GroupRequiredMixin,View):
    template_name = "Ferrocarril/solicitudes_transporte.html"
    group_required = 'AgenteAduanal'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        contenedores_seleccionados = self.request.session.get('contenedores_seleccionados', [])
        bl = self.request.session.get('bl', '')
        
        if not contenedores_seleccionados:
            # messages.warning(self.request, 'No hay contenedores seleccionados')
            return redirect('ferrocarril:ferrocarril_transporte_escoger_contenedor')
        
        context.update({
            'contenedores_seleccionados': contenedores_seleccionados,
            'bl': bl,
            'cantidad_contenedores': len(contenedores_seleccionados),
            'formset_contenedores': zip(context.get('formset', []), contenedores_seleccionados),
        })
        return context
    
    def get(self, request):
        contenedores_seleccionados = self.request.session.get('contenedores_seleccionados', [])

        ContenedorFormSet = modelformset_factory(
            SoliciudIndividualTransporte,
            form = CrearSolicitudesTransporteForm,
            extra=len(contenedores_seleccionados)
        )
        formset = ContenedorFormSet(queryset=SoliciudIndividualTransporte.objects.none(),form_kwargs={'usuario': request.user})
        return render(request, 'Ferrocarril/solicitudes_transporte.html', {
            'formset': formset,
            'formset_contenedores': zip(formset, contenedores_seleccionados)
        })

    def post(self, request):
        contenedores_seleccionados = self.request.session.get('contenedores_seleccionados', [])
        ContenedorFormSet = modelformset_factory(
            SoliciudIndividualTransporte,
            form = CrearSolicitudesTransporteForm,
            extra=len(contenedores_seleccionados)
        )
        formset = ContenedorFormSet(request.POST, form_kwargs={'usuario': request.user})
        if formset.is_valid():
            formset.save()
            return redirect('success_url')
        return render(request, 'Ferrocarril/solicitudes_transporte.html', {'formset': formset})
